#!/usr/bin/env python3
"""Compare NACA 0012 Airfoil 360 measurements with NACA Airfoil Kit's current model.

Source dataset: Stringer, D. Blake (2022), Airfoil 360 v2022: Wind Tunnel
Data, Mendeley Data, V1, DOI 10.17632/dz4bv26ncd.1 (CC BY 4.0).

The script produces a PNG overlay, a row-level residual CSV and a JSON metrics
file. It intentionally labels the package's lightweight panel/empirical result
as a preliminary model; it is not an XFOIL or experimental result.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import load_workbook

from airfoil_pro import ExperimentalValidation, NACAGeneratorPro


SOURCE_URL = "https://data.mendeley.com/datasets/dz4bv26ncd/1"


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True, help="Downloaded Airfoil360_wind_tunnel_data_v2022.xlsx path")
    parser.add_argument("--output-dir", type=Path, default=Path("analysis_outputs/naca0012_airfoil360"))
    parser.add_argument("--alpha-min", type=float, default=0.0, help="Minimum measured alpha included in validation")
    parser.add_argument("--alpha-max", type=float, default=8.1, help="Maximum measured alpha included in validation")
    parser.add_argument("--reynolds", type=int, nargs="+", default=[50_000, 100_000], choices=[50_000, 100_000])
    return parser.parse_args()


def load_experimental_rows(workbook_path: Path, reynolds: int, alpha_min: float, alpha_max: float) -> list[dict[str, float]]:
    sheet_name = f"NACA 0012 Re={reynolds // 1000}K"
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet {sheet_name!r} is not present in {workbook_path}.")
    rows: list[dict[str, float]] = []
    for alpha, cl, cd in workbook[sheet_name].iter_rows(min_row=3, values_only=True):
        if alpha is None or cl is None or cd is None:
            continue
        alpha, cl, cd = float(alpha), float(cl), float(cd)
        if alpha_min <= alpha <= alpha_max:
            rows.append({"alpha_deg": alpha, "cl": cl, "cd": cd})
    if len(rows) < 2:
        raise ValueError(f"{sheet_name} has fewer than two rows in the selected alpha range.")
    return rows


def compare_case(reynolds: int, experimental_rows: list[dict[str, float]]):
    coords = NACAGeneratorPro.naca4("0012", 100)
    if coords is None:
        raise RuntimeError("NACA 0012 geometry could not be generated.")
    comparison = ExperimentalValidation.compare_polar(*coords, experimental_rows, re=reynolds, rough=0.0)
    for row in comparison["comparison"]:
        row["reynolds"] = reynolds
        row["source"] = "Airfoil 360 v2022 experimental wind-tunnel data"
        row["model"] = "NACA Airfoil Kit preliminary panel/empirical"
    return comparison


def write_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = [
        "reynolds", "source", "model", "alpha_deg", "experimental_cl", "model_cl", "cl_error",
        "experimental_cd", "model_cd", "cd_error", "stalled_estimate",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def render_overlay(path: Path, cases: list[dict], alpha_min: float, alpha_max: float) -> None:
    figure, axes = plt.subplots(len(cases), 2, figsize=(12, 4.6 * len(cases)), squeeze=False)
    for row_index, case in enumerate(cases):
        rows = case["comparison"]
        alpha = [row["alpha_deg"] for row in rows]
        for column, metric, label in ((0, "cl", "$C_l$"), (1, "cd", "$C_d$")):
            axis = axes[row_index][column]
            axis.scatter(alpha, [row[f"experimental_{metric}"] for row in rows], color="#f97316", label="Airfoil 360 experiment", zorder=3)
            axis.plot(alpha, [row[f"model_{metric}"] for row in rows], color="#0ea5e9", linewidth=2.2, label="Current preliminary model")
            axis.set_title(f"NACA 0012 | Re = {case['reynolds']:,} | {label}")
            axis.set_xlabel("Angle of attack, α [deg]")
            axis.set_ylabel(label)
            axis.set_xlim(alpha_min - 0.5, alpha_max + 0.5)
            axis.grid(True, alpha=0.3)
            axis.legend(loc="best")
    figure.suptitle("NACA 0012: Airfoil 360 experiment vs current preliminary model", fontsize=14, y=0.995)
    figure.text(0.5, 0.002, "Measured data: Airfoil 360 v2022 (CC BY 4.0). Numerical comparison is not a validation claim outside this range.", ha="center", fontsize=9)
    figure.tight_layout(rect=(0, 0.04, 1, 0.97))
    figure.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(figure)


def main() -> None:
    args = parse_arguments()
    if not args.workbook.is_file():
        raise FileNotFoundError(f"Workbook not found: {args.workbook}")
    if args.alpha_max <= args.alpha_min:
        raise ValueError("--alpha-max must be greater than --alpha-min.")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    cases, all_rows, metrics = [], [], []
    for reynolds in args.reynolds:
        experimental_rows = load_experimental_rows(args.workbook, reynolds, args.alpha_min, args.alpha_max)
        result = compare_case(reynolds, experimental_rows)
        cases.append({"reynolds": reynolds, "comparison": result["comparison"]})
        all_rows.extend(result["comparison"])
        metrics.append(
            {
                "reynolds": reynolds,
                "alpha_min_deg": args.alpha_min,
                "alpha_max_deg": args.alpha_max,
                "point_count": len(result["comparison"]),
                "cl_metrics": result["cl_metrics"],
                "cd_metrics": result["cd_metrics"],
                "source_url": SOURCE_URL,
                "model_scope": "preliminary panel/empirical; not a viscous solver",
            }
        )

    residual_path = args.output_dir / "naca0012_airfoil360_residuals.csv"
    metrics_path = args.output_dir / "naca0012_airfoil360_metrics.json"
    chart_path = args.output_dir / "naca0012_airfoil360_comparison.png"
    write_csv(residual_path, all_rows)
    metrics_path.write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    render_overlay(chart_path, cases, args.alpha_min, args.alpha_max)

    print(f"Wrote {residual_path}")
    print(f"Wrote {metrics_path}")
    print(f"Wrote {chart_path}")
    for item in metrics:
        print(
            f"Re={item['reynolds']:,} | n={item['point_count']} | "
            f"Cl RMSE={item['cl_metrics']['rmse']:.5f} | Cd RMSE={item['cd_metrics']['rmse']:.5f}"
        )


if __name__ == "__main__":
    main()
